from classes.aplicacao import Configuracao
from classes.proposicao import Proposicao
import openpyxl

class PlanilhaProjetos:
    def __init__(self, ordem_inicial):
        self.config = Configuracao()
        self._ordem_inicial = int(ordem_inicial)

    def CarregaColunas(self):
        projetos_selecionados = []
        workbook = openpyxl.load_workbook(self.config.arquivo_planilha_de_projetos, keep_vba=True)
        planilha = workbook[self.config.planilha_de_projetos]

        conta = 0
        for linha in planilha.iter_rows():
            if linha[2].value == None and conta > 1:
                break

            if linha[self.config.coluna_reuniao-1].value == self.config.filtro_coluna_reuniao:
                conta += 1
                #print(f"col0: {linha[self.config.coluna_tipo_projeto-1].value} \ncol1: {linha[self.config.coluna_numero_projeto-1].value} \ncol2: {linha[self.config.coluna_autor-1].value} \n col3: {linha[3].value}\n col4: {linha[4].value}")
                #print()  # Passa para a próxima linha após imprimir todas as células da linha atual

                proposicao = Proposicao()
                proposicao.tipo_proposicao = str(linha[self.config.coluna_tipo_projeto-1].value).strip()
                proposicao.numero = str(linha[self.config.coluna_numero_projeto-1].value).strip()
                proposicao.ementa = linha[self.config.coluna_ementa-1].value.strip()
                proposicao.autores = linha[self.config.coluna_autor-1].value.strip().upper()
                proposicao.parecer = linha[self.config.coluna_parecer-1].value.strip().upper()
                proposicao.relator = linha[self.config.coluna_relatoria-1].value.strip().upper()
                proposicao.reuniao = linha[self.config.coluna_reuniao-1].value.strip()
                proposicao.linha_da_planilha = linha[0].row  #ARMAZENO A LINHA PARA DEPOIS ATUALIZAR A ORDEM NO EDITAL

                if "(EP)" in proposicao.numero:
                    proposicao.numero = proposicao.numero.replace("(EP)", "").strip(" ")
                    proposicao.emenda_de_plenario = True
                else:
                    if "EP" in proposicao.numero:
                        proposicao.numero = proposicao.numero.replace("EP", "").strip(" ")
                        proposicao.emenda_de_plenario = True

                num = proposicao.numero.split('/')
                proposicao.numero = num[0]
                proposicao.ano = num[-1]


                projetos_selecionados.append(proposicao)

        # Ordena as proposições
        if len(projetos_selecionados) > 0:
            ##### projetos_selecionados = sorted(projetos_selecionados, key=lambda x: (x.relator, int(x.ano), int(x.numero)))

            presidente_comissao = self.config.presidente_comissao  # nome do relator/presidente da comissão que deve ficar no topo

            projetos_selecionados = sorted(
                projetos_selecionados,
                key=lambda x: (
                    0 if (x.relator or "").strip().upper() == presidente_comissao.upper() else 1,  # prioridade
                    (x.relator or "").strip().upper(),  # relator normal
                    int(x.ano) if str(x.ano).isdigit() else 0,
                    int(x.numero) if str(x.numero).isdigit() else 0
                )
            )

            indice = 0
            while indice < len(projetos_selecionados):
                projetos_selecionados[indice].ordem = self._ordem_inicial
                indice += 1
                self._ordem_inicial += 1

        workbook.close()

        return projetos_selecionados


    def AtualizaOrdemNosProjetos(self, proposicoes):
        try:
            workbook = openpyxl.load_workbook(self.config.arquivo_planilha_de_projetos, keep_vba=True)
            planilha = workbook[self.config.planilha_de_projetos]

            for proposicao in proposicoes:
                linha_idx = proposicao.linha_da_planilha

                # Ignora se a linha não foi definida
                if not linha_idx or linha_idx <= 0:
                    continue

                # Lê as células relevantes
                celula_ordem = planilha.cell(
                    row=linha_idx, column=self.config.coluna_ordem
                )

                celula_ordem.value = proposicao.ordem


            workbook.save(self.config.arquivo_planilha_de_projetos)
            workbook.close()

            return "OK"

        except Exception as e:
            return f'ERRO: {e}'




'''
Testes de execução do módulo.
'''
if __name__ == '__main__':

    print("Teste de execução")
    P = PlanilhaProjetos()
    print("Carregando projetos")
    p = P.CarregaColunas()
    print(f"Projetos selecionados: {len(p)}")
